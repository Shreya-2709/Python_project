import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font  


def ensure_folder_exists(folder_name):
    if not os.path.exists(folder_name):
        os.makedirs(folder_name)


def save_to_excel(data, excel_folder, filename):
    ensure_folder_exists(excel_folder)  
    excel_path = os.path.join(excel_folder, filename)
    rows = []
    for page, key_value_pairs in data.items():
        rows.append({"Key": "", "Value": ""})  
        rows.append({"Key": "Page", "Value": page})  
        for key, value in key_value_pairs.items():
            rows.append({"Key": key, "Value": value})

    df = pd.DataFrame(rows, columns=["Key", "Value"])  
    df.to_excel(excel_path, index=False)
    wb = load_workbook(excel_path)  
    ws = wb.active  
    for row in ws.iter_rows():
        if row[0].value == "Page":  
            for cell in row:
                cell.font = Font(bold=True)  

    wb.save(excel_path)

    return excel_path
