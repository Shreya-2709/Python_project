import streamlit as st
import os
import re
import pytesseract
import cv2
import pandas as pd


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter  
from PIL import Image
from enhancement import enhance_image
from image_conversion import convert_pdf_to_images
from key_value_extraction import extract_key_value_pairs
from clear_output import clear_output_folder
from excel_operations import save_to_excel, ensure_folder_exists


def adjust_column_width(excel_path):
    workbook = load_workbook(excel_path)
    worksheet = workbook.active

    for column in worksheet.columns:
        max_length = 0
        column = list(column)  
        column_letter = get_column_letter(column[0].column) 
        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

        worksheet.column_dimensions[column_letter].width = max_length + 2

    workbook.save(excel_path)


def main():
    st.title("PDF to JPG Converter and Key-Value Extraction")

    # File uploader for PDFs
    pdf_file = st.file_uploader("Upload a PDF", type=["pdf"])

    if pdf_file:
        pdf_path = "uploaded_pdf.pdf"
        with open(pdf_path, "wb") as f:
            f.write(pdf_file.getbuffer())

        # Convert the PDF to JPG
        output_folder = "pdf_images"
        ensure_folder_exists(output_folder)
        image_paths = convert_pdf_to_images(pdf_path, output_folder)

        st.write(f"Converted {len(image_paths)} pages from the PDF.")

        # "Select All" checkbox
        if "select_all" not in st.session_state:
            st.session_state["select_all"] = False
        
        selected_checkboxes = []
        select_all = st.checkbox("Select All", key="select_all") 

        # Checkboxes for selecting pages
        for i, image_path in enumerate(image_paths):
            page_label = f"Page {i + 1}"
            is_checked = st.checkbox(page_label, key=f"page_{i}", value=select_all)
            if is_checked:
                selected_checkboxes.append((page_label, image_path))

        # Extract key-value pairs from selected images
        if st.button("Extract the key-value Pairs"):
            enhanced_folder = "enhanced_images"
            ensure_folder_exists(enhanced_folder)

            all_key_value_pairs = {}

            for page_label, image_path in selected_checkboxes:
                # Enhance the image before OCR
                enhanced_image = enhance_image(image_path)
                enhanced_image_path = os.path.join(enhanced_folder, os.path.basename(image_path))
                cv2.imwrite(enhanced_image_path, enhanced_image)

                # Extract text with OCR and clean it
                text = re.sub(r'^[\s_-]+$', '', pytesseract.image_to_string(enhanced_image_path), flags=re.MULTILINE)

                # Extract key-value pairs
                key_value_pairs = extract_key_value_pairs(text)

                all_key_value_pairs[page_label] = key_value_pairs

                # Display images and key-value pairs in parallel
                col1, col2 = st.columns(2)
                with col1:
                    st.image(enhanced_image_path, caption=f"Enhanced {os.path.basename(image_path)}", use_column_width=True)
                with col2:
                    for key, value in key_value_pairs.items():
                        st.write(f"**{key}:** {value}")
                        st.markdown("---")

            # Save key-value pairs to Excel 
            excel_folder = "excel_files"
            ensure_folder_exists(excel_folder)
            excel_path = save_to_excel(all_key_value_pairs, excel_folder, "extracted_data.xlsx")

            # Adjust the column width to fit content
            adjust_column_width(excel_path)

            # Success message after Excel sheet is created
            st.success("Excel sheet created successfully.")

            # Download button for the Excel file
            with open(excel_path, "rb") as f:
                excel_data = f.read()

            st.download_button(
                label="Download Excel",
                data=excel_data,
                file_name="extracted_data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        os.remove(pdf_path)

    # Button to clear cached images
    if st.button("Clear Cached Images"):
        clear_output_folder("pdf_images")
        st.success("Cached images cleared.")


if __name__ == "__main__":
    main()
